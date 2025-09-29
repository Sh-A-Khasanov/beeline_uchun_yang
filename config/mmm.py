import json
import openpyxl

# JSON faylni o‘qish
with open("config/num.json", "r", encoding="utf-8") as f:
    data = json.load(f)  # Bu holda data list bo‘ladi

# Excel fayl yaratish
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Numbers"

# Sarlavha yozish
ws["A1"] = "A (kod)"
ws["B1"] = "B"
ws["C1"] = "C"
ws["D1"] = "D"

row = 2
for item in data:
    phone = item.get("phoneNumber", "")
    if phone.startswith("998") and len(phone) == 12:
        # 998906638008 -> 90 | 663 | 80 | 08
        kod = phone[3:5]
        part1 = phone[5:8]
        part2 = phone[8:10]
        part3 = phone[10:12]

        ws.cell(row=row, column=1, value=kod)
        ws.cell(row=row, column=2, value=part1)
        ws.cell(row=row, column=3, value=part2)
        ws.cell(row=row, column=4, value=part3)
        row += 1

# Faylni saqlash
wb.save("numbers.xlsx")
print("Tayyor: numbers.xlsx")
